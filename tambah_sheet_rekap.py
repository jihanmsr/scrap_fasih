import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE = 'muatan_sls_72.xlsx'

# ── 1. Baca data ──────────────────────────────────────────────────────────────
print("Membaca data...")
df = pd.read_excel(FILE, sheet_name='muatan_sls')
df['usaha'] = df['umkm_keluarga'] + df['jml_utp_subsektor'] + df['Total_usaha_SBR']

id_sls    = ['kdprov','nmprov','kdkab','nmkab','kdkec','nmkec','kddesa','nmdesa','kdsls','kdsubsls','nmsls']
id_desa   = ['kdprov','nmprov','kdkab','nmkab','kdkec','nmkec','kddesa','nmdesa']
id_kec    = ['kdprov','nmprov','kdkab','nmkab','kdkec','nmkec']
id_kabkot = ['kdprov','nmprov','kdkab','nmkab']

agg = dict(
    keluarga      = ('keluarga',          'sum'),
    umkm_keluarga = ('umkm_keluarga',     'sum'),
    utp_subsektor = ('jml_utp_subsektor', 'sum'),
    sbr           = ('Total_usaha_SBR',   'sum'),
    usaha         = ('usaha',             'sum'),
)

rekap_sls    = df.groupby(id_sls,    as_index=False).agg(**agg)
rekap_desa   = df.groupby(id_desa,   as_index=False).agg(**agg)
rekap_kec    = df.groupby(id_kec,    as_index=False).agg(**agg)
rekap_kabkot = df.groupby(id_kabkot, as_index=False).agg(**agg)

print(f"  Per SLS    : {len(rekap_sls):,} baris")
print(f"  Per Desa   : {len(rekap_desa):,} baris")
print(f"  Per Kec    : {len(rekap_kec):,} baris")
print(f"  Per KabKot : {len(rekap_kabkot):,} baris")

# ── 2. Style helpers ──────────────────────────────────────────────────────────
FILL_ORIG    = PatternFill("solid", fgColor="17375E")   # navy (muatan_sls)
FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")   # biru tua
FILL_FORMULA = PatternFill("solid", fgColor="E2EFDA")   # hijau muda
FILL_TOTAL   = PatternFill("solid", fgColor="FFF2CC")   # kuning
HDR_FONT     = Font(bold=True, color="FFFFFF", size=10)
NORM_FONT    = Font(size=10)
BOLD_FONT    = Font(bold=True, size=10)
CENTER       = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT         = Alignment(horizontal='left',   vertical='center')
RIGHT        = Alignment(horizontal='right',  vertical='center')

thin   = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell, fill=None):
    cell.fill      = fill or FILL_HEADER
    cell.font      = HDR_FONT
    cell.alignment = CENTER
    cell.border    = BORDER

def style_data(cell, bold=False, fill=None, align=RIGHT):
    cell.font      = BOLD_FONT if bold else NORM_FONT
    cell.alignment = align
    cell.border    = BORDER
    if fill:
        cell.fill = fill

# ── 3. Buat workbook baru ─────────────────────────────────────────────────────
print("Membangun workbook baru...")
wb = Workbook()

# ── 4. Sheet muatan_sls (data asli) ──────────────────────────────────────────
print("  Menulis sheet muatan_sls...")
ws0 = wb.active
ws0.title = 'muatan_sls'

orig_cols = df.drop(columns=['usaha']).columns.tolist()
ws0.append(orig_cols)
for c in range(1, len(orig_cols)+1):
    style_header(ws0.cell(1, c), fill=FILL_ORIG)

for row_vals in df[orig_cols].itertuples(index=False):
    ws0.append(list(row_vals))

ws0.freeze_panes = 'A2'
ws0.row_dimensions[1].height = 25
print(f"  muatan_sls: {ws0.max_row-1:,} baris")

# ── 5. Sheet Rekap_SLS ────────────────────────────────────────────────────────
print("  Menulis sheet Rekap_SLS...")
ws1 = wb.create_sheet('Rekap_SLS')

headers_sls = [
    'Kode Prov','Nama Prov','Kode Kab','Nama Kab/Kota',
    'Kode Kec','Nama Kec','Kode Desa','Nama Desa',
    'Kode SLS','Kode SubSLS','Nama SLS',
    'Keluarga','UMKM Keluarga','UTP Subsektor','SBR',
    'Usaha\n(=UMKM+UTP+SBR)'
]
ws1.append(headers_sls)
for c in range(1, len(headers_sls)+1):
    style_header(ws1.cell(1, c))

for i, row in rekap_sls.iterrows():
    r = i + 2
    vals = [
        row.kdprov, row.nmprov, row.kdkab, row.nmkab,
        row.kdkec,  row.nmkec,  row.kddesa, row.nmdesa,
        row.kdsls,  row.kdsubsls, row.nmsls,
        row.keluarga, row.umkm_keluarga, row.utp_subsektor, row.sbr,
    ]
    for c, v in enumerate(vals, 1):
        ws1.cell(r, c).value = v
        style_data(ws1.cell(r, c), align=RIGHT if c > 11 else LEFT)

    cell_u = ws1.cell(r, 16)
    cell_u.value = f"=M{r}+N{r}+O{r}"
    style_data(cell_u, fill=FILL_FORMULA)

tr = len(rekap_sls) + 2
ws1.cell(tr, 11).value = "TOTAL"
style_data(ws1.cell(tr, 11), bold=True, fill=FILL_TOTAL, align=RIGHT)
for c_idx, col_l in [(12,'L'),(13,'M'),(14,'N'),(15,'O'),(16,'P')]:
    cell = ws1.cell(tr, c_idx)
    cell.value = f"=SUM({col_l}2:{col_l}{tr-1})"
    style_data(cell, bold=True, fill=FILL_TOTAL)

col_w_sls = [8,16,8,22,8,18,8,24,10,12,28,12,15,15,12,18]
for i, w in enumerate(col_w_sls, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = 'A2'
ws1.row_dimensions[1].height = 35

# ── 6. Sheet Rekap_Desa ───────────────────────────────────────────────────────
print("  Menulis sheet Rekap_Desa...")
ws_desa = wb.create_sheet('Rekap_Desa')

headers_desa = [
    'Kode Prov','Nama Prov','Kode Kab','Nama Kab/Kota',
    'Kode Kec','Nama Kec','Kode Desa','Nama Desa',
    'Keluarga','UMKM Keluarga','UTP Subsektor','SBR',
    'Usaha\n(=UMKM+UTP+SBR)'
]
ws_desa.append(headers_desa)
for c in range(1, len(headers_desa)+1):
    style_header(ws_desa.cell(1, c))

for i, row in rekap_desa.iterrows():
    r = i + 2
    vals = [
        row.kdprov, row.nmprov, row.kdkab, row.nmkab,
        row.kdkec,  row.nmkec,  row.kddesa, row.nmdesa,
        row.keluarga, row.umkm_keluarga, row.utp_subsektor, row.sbr,
    ]
    for c, v in enumerate(vals, 1):
        ws_desa.cell(r, c).value = v
        style_data(ws_desa.cell(r, c), align=RIGHT if c > 8 else LEFT)

    cell_u = ws_desa.cell(r, 13)
    cell_u.value = f"=J{r}+K{r}+L{r}"
    style_data(cell_u, fill=FILL_FORMULA)

tr_desa = len(rekap_desa) + 2
ws_desa.cell(tr_desa, 8).value = "TOTAL"
style_data(ws_desa.cell(tr_desa, 8), bold=True, fill=FILL_TOTAL, align=RIGHT)
for c_idx, col_l in [(9,'I'),(10,'J'),(11,'K'),(12,'L'),(13,'M')]:
    cell = ws_desa.cell(tr_desa, c_idx)
    cell.value = f"=SUM({col_l}2:{col_l}{tr_desa-1})"
    style_data(cell, bold=True, fill=FILL_TOTAL)

col_w_desa = [8,16,8,22,8,18,8,24,12,15,15,12,18]
for i, w in enumerate(col_w_desa, 1):
    ws_desa.column_dimensions[get_column_letter(i)].width = w
ws_desa.freeze_panes = 'A2'
ws_desa.row_dimensions[1].height = 35

# ── 7. Sheet Rekap_Kecamatan ──────────────────────────────────────────────────
print("  Menulis sheet Rekap_Kecamatan...")
ws_kec = wb.create_sheet('Rekap_Kecamatan')

headers_kec = [
    'Kode Prov','Nama Prov','Kode Kab','Nama Kab/Kota',
    'Kode Kec','Nama Kec',
    'Keluarga','UMKM Keluarga','UTP Subsektor','SBR',
    'Usaha\n(=UMKM+UTP+SBR)'
]
ws_kec.append(headers_kec)
for c in range(1, len(headers_kec)+1):
    style_header(ws_kec.cell(1, c))

for i, row in rekap_kec.iterrows():
    r = i + 2
    vals = [
        row.kdprov, row.nmprov, row.kdkab, row.nmkab,
        row.kdkec,  row.nmkec,
        row.keluarga, row.umkm_keluarga, row.utp_subsektor, row.sbr,
    ]
    for c, v in enumerate(vals, 1):
        ws_kec.cell(r, c).value = v
        style_data(ws_kec.cell(r, c), align=RIGHT if c > 6 else LEFT)

    cell_u = ws_kec.cell(r, 11)
    cell_u.value = f"=H{r}+I{r}+J{r}"
    style_data(cell_u, fill=FILL_FORMULA)

tr_kec = len(rekap_kec) + 2
ws_kec.cell(tr_kec, 6).value = "TOTAL"
style_data(ws_kec.cell(tr_kec, 6), bold=True, fill=FILL_TOTAL, align=RIGHT)
for c_idx, col_l in [(7,'G'),(8,'H'),(9,'I'),(10,'J'),(11,'K')]:
    cell = ws_kec.cell(tr_kec, c_idx)
    cell.value = f"=SUM({col_l}2:{col_l}{tr_kec-1})"
    style_data(cell, bold=True, fill=FILL_TOTAL)

col_w_kec = [8,16,8,22,8,18,12,15,15,12,18]
for i, w in enumerate(col_w_kec, 1):
    ws_kec.column_dimensions[get_column_letter(i)].width = w
ws_kec.freeze_panes = 'A2'
ws_kec.row_dimensions[1].height = 35

# ── 8. Sheet Rekap_KabKot ─────────────────────────────────────────────────────
print("  Menulis sheet Rekap_KabKot...")
ws2 = wb.create_sheet('Rekap_KabKot')

headers_kk = [
    'Kode Prov','Nama Prov','Kode Kab','Nama Kab/Kota',
    'Keluarga', 'UMKM Keluarga', 'UTP Subsektor', 'SBR',
    'Usaha\n(=UMKM+UTP+SBR)'
]
ws2.append(headers_kk)
for c in range(1, len(headers_kk)+1):
    style_header(ws2.cell(1, c))

for i, row in rekap_kabkot.iterrows():
    r = i + 2
    ws2.cell(r,1).value = row.kdprov;  style_data(ws2.cell(r,1), align=CENTER)
    ws2.cell(r,2).value = row.nmprov;  style_data(ws2.cell(r,2), align=LEFT)
    ws2.cell(r,3).value = row.kdkab;   style_data(ws2.cell(r,3), align=CENTER)
    ws2.cell(r,4).value = row.nmkab;   style_data(ws2.cell(r,4), align=LEFT)

    ws2.cell(r,5).value = row.keluarga;       style_data(ws2.cell(r,5), align=RIGHT)
    ws2.cell(r,6).value = row.umkm_keluarga;  style_data(ws2.cell(r,6), align=RIGHT)
    ws2.cell(r,7).value = row.utp_subsektor;  style_data(ws2.cell(r,7), align=RIGHT)
    ws2.cell(r,8).value = row.sbr;            style_data(ws2.cell(r,8), align=RIGHT)

    ws2.cell(r,9).value = f"=E{r}+F{r}+G{r}+H{r}"
    style_data(ws2.cell(r,9), fill=FILL_FORMULA)

# Total row
tr2 = len(rekap_kabkot) + 2
ws2.cell(tr2,4).value = "TOTAL"
style_data(ws2.cell(tr2,4), bold=True, fill=FILL_TOTAL, align=RIGHT)
for c_idx, col_l in [(5,'E'),(6,'F'),(7,'G'),(8,'H'),(9,'I')]:
    cell = ws2.cell(tr2, c_idx)
    cell.value = f"=SUM({col_l}2:{col_l}{tr2-1})"
    style_data(cell, bold=True, fill=FILL_TOTAL)

col_w_kk = [8,16,8,24,16,18,18,16,20]
for i, w in enumerate(col_w_kk, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'
ws2.row_dimensions[1].height = 40

# ── 9. Simpan ─────────────────────────────────────────────────────────────────
print("Menyimpan file...")
wb.save(FILE)
print(f"\n✅ Selesai! File disimpan: {FILE}")
print(f"   Sheets: {wb.sheetnames}")
