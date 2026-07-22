import openpyxl
from openpyxl.utils import get_column_letter

file_name = 'hasil_bpom_prov_72_full_final.xlsx'
wb = openpyxl.load_workbook(file_name)

if 'Rekapitulasi' in wb.sheetnames:
    del wb['Rekapitulasi']
ws = wb.create_sheet('Rekapitulasi', 0)

kab_map = {
    7201: 'BANGGAI KEPULAUAN', 7202: 'BANGGAI', 7203: 'MOROWALI', 7204: 'POSO',
    7205: 'DONGGALA', 7207: 'BUOL', 7208: 'PARIGI MOUTONG',
    7209: 'TOJO UNA-UNA', 7210: 'SIGI', 7211: 'BANGGAI LAUT', 7212: 'MOROWALI UTARA',
    7271: 'PALU'
}

# Header Row 1
ws.merge_cells('A1:A2')
ws['A1'] = 'Kode Kabupaten'
ws.merge_cells('B1:D1')
ws['B1'] = 'Keberadaan Usaha'
ws.merge_cells('E1:E2')
ws['E1'] = 'Jumlah'
ws.merge_cells('F1:F2')
ws['F1'] = 'Tindak Lanjut'
ws.merge_cells('G1:G2')
ws['G1'] = 'Persentase'

# Header Row 2
ws['B2'] = '0. Tidak Ditemukan'
ws['C2'] = '1. Ditemukan'
ws['D2'] = '2. Baru'

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15

row_num = 3
for kdkab in sorted(kab_map.keys()):
    ws.cell(row=row_num, column=1).value = kdkab
    
    # Kdkab in data is the last 2 digits, so we use VALUE(RIGHT(A3,2)) to get 1, 2, 71 etc.
    # We can also just put the true kdkab value directly to avoid complex formulas, but let's use the formula as asked
    real_kdkab = int(str(kdkab)[2:])
    
    # 0. Tidak Ditemukan (Column B, idx 2) -> Sheet1!R:R is Keberadaan Usaha Value
    ws.cell(row=row_num, column=2).value = f'=COUNTIFS(Sheet1!C:C, {real_kdkab}, Sheet1!R:R, 0)'
    
    # 1. Ditemukan (Column C, idx 3)
    ws.cell(row=row_num, column=3).value = f'=COUNTIFS(Sheet1!C:C, {real_kdkab}, Sheet1!R:R, 1)'
    
    # 2. Baru (Column D, idx 4)
    ws.cell(row=row_num, column=4).value = f'=COUNTIFS(Sheet1!C:C, {real_kdkab}, Sheet1!R:R, 2)'
    
    # Jumlah (Column E, idx 5)
    ws.cell(row=row_num, column=5).value = f'=SUM(B{row_num}:D{row_num})'
    
    # Tindak Lanjut (Column F, idx 6) -> Sheet1!X:X is Hasil Tindak Lanjut. Count not empty.
    ws.cell(row=row_num, column=6).value = f'=COUNTIFS(Sheet1!C:C, {real_kdkab}, Sheet1!X:X, "<>")'
    
    # Persentase (Column G, idx 7)
    ws.cell(row=row_num, column=7).value = f'=IF(E{row_num}>0, (F{row_num}/E{row_num})*100, 0)'
    
    row_num += 1

# Total row
ws.cell(row=row_num, column=1).value = 7200
ws.cell(row=row_num, column=2).value = f'=SUM(B3:B{row_num-1})'
ws.cell(row=row_num, column=3).value = f'=SUM(C3:C{row_num-1})'
ws.cell(row=row_num, column=4).value = f'=SUM(D3:D{row_num-1})'
ws.cell(row=row_num, column=5).value = f'=SUM(E3:E{row_num-1})'
ws.cell(row=row_num, column=6).value = f'=SUM(F3:F{row_num-1})'
ws.cell(row=row_num, column=7).value = f'=IF(E{row_num}>0, (F{row_num}/E{row_num})*100, 0)'

wb.save(file_name)
print("Updated Rekapitulasi to match requested format")
