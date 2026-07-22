import openpyxl
from openpyxl.utils import get_column_letter

file_name = 'hasil_bpom_prov_72_full_final.xlsx'
wb = openpyxl.load_workbook(file_name)

# Create or get sheet
if 'Rekapitulasi' in wb.sheetnames:
    del wb['Rekapitulasi']
ws_rekap = wb.create_sheet('Rekapitulasi', 0) # insert as first sheet

# Kabupaten Map
kab_map = {
    1: 'BANGGAI KEPULAUAN', 2: 'BANGGAI', 3: 'MOROWALI', 4: 'POSO',
    5: 'DONGGALA', 6: 'TOLI-TOLI', 7: 'BUOL', 8: 'PARIGI MOUTONG',
    9: 'TOJO UNA-UNA', 10: 'SIGI', 11: 'BANGGAI LAUT', 12: 'MOROWALI UTARA',
    71: 'PALU'
}

# Headers
headers = [
    "Kode Kab", "Nama Kabupaten", "Total Usaha (Target)",
    "APPROVED BY Pengawas", "SUBMITTED BY Pencacah",
    "REJECTED BY Pengawas", "REVOKED BY Pengawas", "EDITED BY Admin Kabupaten", "NOT_FOUND / ERROR"
]

for col_num, header in enumerate(headers, 1):
    cell = ws_rekap.cell(row=1, column=col_num)
    cell.value = header
    ws_rekap.column_dimensions[get_column_letter(col_num)].width = 25

# Data rows (formulas)
row_num = 2
for kdkab, nama in sorted(kab_map.items()):
    ws_rekap.cell(row=row_num, column=1).value = kdkab
    ws_rekap.cell(row=row_num, column=2).value = nama
    
    # Formula Total: =COUNTIF(Sheet1!C:C, A2)
    ws_rekap.cell(row=row_num, column=3).value = f'=COUNTIF(Sheet1!C:C, A{row_num})'
    
    # Formula APPROVED: =COUNTIFS(Sheet1!C:C, A2, Sheet1!AC:AC, "APPROVED BY Pengawas")
    ws_rekap.cell(row=row_num, column=4).value = f'=COUNTIFS(Sheet1!C:C, A{row_num}, Sheet1!AC:AC, "APPROVED BY Pengawas")'
    
    # Formula SUBMITTED: =COUNTIFS(Sheet1!C:C, A2, Sheet1!AC:AC, "SUBMITTED BY Pencacah")
    ws_rekap.cell(row=row_num, column=5).value = f'=COUNTIFS(Sheet1!C:C, A{row_num}, Sheet1!AC:AC, "SUBMITTED BY Pencacah")'
    
    # Formula REJECTED: =COUNTIFS(Sheet1!C:C, A2, Sheet1!AC:AC, "REJECTED BY Pengawas")
    ws_rekap.cell(row=row_num, column=6).value = f'=COUNTIFS(Sheet1!C:C, A{row_num}, Sheet1!AC:AC, "REJECTED BY Pengawas")'
    
    # Formula REVOKED: =COUNTIFS(Sheet1!C:C, A2, Sheet1!AC:AC, "REVOKED BY Pengawas")
    ws_rekap.cell(row=row_num, column=7).value = f'=COUNTIFS(Sheet1!C:C, A{row_num}, Sheet1!AC:AC, "REVOKED BY Pengawas")'
    
    # Formula EDITED ADMIN: =COUNTIFS(Sheet1!C:C, A2, Sheet1!AC:AC, "EDITED BY Admin Kabupaten")
    ws_rekap.cell(row=row_num, column=8).value = f'=COUNTIFS(Sheet1!C:C, A{row_num}, Sheet1!AC:AC, "EDITED BY Admin Kabupaten")'
    
    # Formula NOT_FOUND: =COUNTIFS(Sheet1!C:C, A2, Sheet1!AC:AC, "NOT_FOUND") + COUNTIFS(Sheet1!C:C, A2, Sheet1!AC:AC, "ERROR*")
    ws_rekap.cell(row=row_num, column=9).value = f'=COUNTIFS(Sheet1!C:C, A{row_num}, Sheet1!AC:AC, "NOT_FOUND") + COUNTIFS(Sheet1!C:C, A{row_num}, Sheet1!AC:AC, "ERROR*")'
    
    row_num += 1

# Optional: Add total row at the bottom
ws_rekap.cell(row=row_num, column=2).value = "TOTAL KESELURUHAN"
for c in range(3, 10):
    ws_rekap.cell(row=row_num, column=c).value = f'=SUM({get_column_letter(c)}2:{get_column_letter(c)}{row_num-1})'

wb.save(file_name)
print("Added rekap formulas!")
