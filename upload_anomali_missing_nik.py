"""
upload_anomali_missing_nik.py
Upload data anomali "Missing NIK Anggota Keluarga" dari file Excel ke Supabase.
Mendukung deduplication (tidak upload ulang yang sudah ada).
"""
import sys
import os
import json
import openpyxl
from datetime import datetime, timezone, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from merge_granulars import load_supabase_config

JENIS_ANOMALI = "Missing NIK Anggota Keluarga"
PERIODE_SE = "fd68e454-ba45-4b85-8205-f3bf777ded24"

# ===================== LOAD PETUGAS DATABASE =====================
def load_petugas_map():
    """Load petugas username→fullname map dari assign_data.js atau local files."""
    pet_map = {}  # email → fullname
    
    # Coba load dari assign_data.js
    for fname in ['assign_data.js', 'ipas_data.js']:
        fpath = os.path.join(os.path.dirname(os.path.abspath(__file__)), fname)
        if os.path.exists(fpath):
            try:
                with open(fpath, 'r', encoding='utf-8') as f:
                    content = f.read()
                # Extract PETUGAS_DATA_UMUM jika ada
                import re
                m = re.search(r'window\.PETUGAS_DATA_UMUM\s*=\s*(\[.*?\])\s*;', content, re.DOTALL)
                if m:
                    petugas_list = json.loads(m.group(1))
                    for p in petugas_list:
                        email = p.get('username', '').strip().lower()
                        name = p.get('fullname', '').strip()
                        if email and name:
                            pet_map[email] = name
                    print(f"[INFO] Loaded {len(pet_map)} petugas dari {fname}")
                    break
            except Exception as e:
                print(f"[WARN] Gagal baca {fname}: {e}")
    
    if not pet_map:
        print("[WARN] Petugas map kosong - nama petugas akan diambil langsung dari data Excel")
    
    return pet_map


# ===================== PROCESS EXCEL =====================
def read_missing_nik_excel(filepath):
    print(f"[INFO] Membaca file: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"[INFO] Headers ({len(headers)} kolom): {headers[:10]}...")
    
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = dict(zip(headers, row))
        rows.append(row_dict)
    
    print(f"[INFO] Total {len(rows)} baris data")
    return rows


def build_catatan(row):
    """Gabungkan semua catatan field menjadi satu teks terstruktur."""
    parts = []
    
    # Hubungan & keberadaan
    hubungan = row.get('hubungan_label', '')
    keberadaan = row.get('keberadaan_dtsen_label', '')
    no_kk = row.get('no_urut_kk', '')
    
    if hubungan:
        parts.append(f"Hubungan: {hubungan}")
    if keberadaan:
        parts.append(f"Keberadaan: {keberadaan}")
    if no_kk:
        parts.append(f"No. KK: {no_kk}")
    
    # Alamat
    alamat = row.get('jalan_domisili', '') or row.get('alamat_prelist', '')
    if alamat and str(alamat).strip():
        parts.append(f"Alamat: {str(alamat).strip()}")
    
    # Catatan petugas (dari berbagai field catatan)
    catatan_fields = ['catatan', 'catatan_1', 'catatan_2', 'catatan_3', 'catatan_pml']
    catatan_values = []
    for cf in catatan_fields:
        val = row.get(cf)
        if val and str(val).strip():
            catatan_values.append(str(val).strip())
    
    if catatan_values:
        parts.append("Catatan: " + " | ".join(catatan_values))
    
    # Comment
    comment = row.get('comment', '')
    if comment and str(comment).strip() and str(comment).strip() != '{"dataKey":"","notes":[]}':
        try:
            c = json.loads(str(comment))
            notes = c.get('notes', [])
            if notes:
                parts.append(f"Komentar: {'; '.join(str(n) for n in notes)}")
        except:
            if comment:
                parts.append(f"Komentar: {comment}")
    
    return " | ".join(parts) if parts else ""


def process_rows(rows, petugas_map):
    """Convert Excel rows ke format anomali_data Supabase."""
    records = []
    wita = timezone(timedelta(hours=8))
    now_str = datetime.now(wita).isoformat()
    
    for row in rows:
        # Kab
        kab_code = str(row.get('level_2_name', '') or '').strip()
        kec_code = str(row.get('level_3_code', '') or '').strip()
        desa_code = str(row.get('level_4_code', '') or '').strip()
        sls_code = str(row.get('level_5_code', '') or '').strip()
        
        kab_label = str(row.get('level_2_code', '') or '').strip() + ' - ' + kab_code
        kec_name = str(row.get('level_3_name', '') or '').strip()
        desa_name = str(row.get('level_4_name', '') or '').strip()
        sls_name = str(row.get('level_5_name', '') or '').strip()
        
        # Nama anggota dengan NIK kosong
        nama = str(row.get('nama_dtsen', '') or '').strip()
        nik = str(row.get('nik_dtsen', '9999') or '9999').strip()
        
        # Petugas
        email_petugas = str(row.get('current_user_username', '') or '').strip().lower()
        nama_petugas = str(row.get('current_user_fullname', '') or '').strip()
        role_petugas = str(row.get('current_user_survey_role_name', '') or '').strip()
        
        # Enrich nama_petugas dari map jika kosong atau "-"
        if (not nama_petugas or nama_petugas == '-') and email_petugas:
            nama_petugas = petugas_map.get(email_petugas, nama_petugas or '-')
        
        # Tambahkan role ke nama petugas
        if role_petugas and nama_petugas and nama_petugas != '-':
            nama_petugas_display = f"{nama_petugas} ({role_petugas})"
        else:
            nama_petugas_display = nama_petugas or '-'
        
        # catatan gabungan
        catatan = build_catatan(row)
        
        # assignment
        assignment_id = str(row.get('assignment_id', '') or '').strip()
        link = f"https://fasih-sm.bps.go.id/app/assignment/{PERIODE_SE}/{assignment_id}" if assignment_id else ''
        
        record = {
            'kab_code': kab_code,           # Nama kab (BANGGAI, PALU, dll)
            'kec_code': kec_name,           # Nama kec (bukan code, untuk display)
            'desa_code': desa_name,         # Nama desa
            'sls_code': sls_name,           # Nama SLS
            'nama_petugas': nama_petugas_display,
            'jenis_anomali': JENIS_ANOMALI,
            'nama_krt': nama,               # Nama anggota dgn NIK 9999
            'catatan': catatan,             # Catatan gabungan
            'tindak_lanjut': '',
            'status_anomali': 1,            # Perlu Tindak Lanjut
            'assignment_id': assignment_id,
            'total_pengeluaran': 0,
            'biaya_produksi': 0,
            'pct_biaya': 0.0,
            'waktu_anomali': now_str,
            # Extra context packed into catatan sudah ada
        }
        records.append(record)
    
    return records


# ===================== UPLOAD =====================
def upload_to_supabase(records, supabase):
    print(f"\n[UPLOAD] Total {len(records)} record akan diupload...")
    
    # Cek yang sudah ada (dedup by assignment_id + nama_krt)
    existing_res = supabase.table('anomali_data')\
        .select('assignment_id, nama_krt')\
        .eq('jenis_anomali', JENIS_ANOMALI)\
        .execute()
    
    existing_keys = set()
    for row in (existing_res.data or []):
        key = f"{row.get('assignment_id', '')}|{row.get('nama_krt', '')}"
        existing_keys.add(key)
    
    print(f"[INFO] Sudah ada {len(existing_keys)} record '{JENIS_ANOMALI}' di DB")
    
    new_records = []
    skip_count = 0
    for rec in records:
        key = f"{rec['assignment_id']}|{rec['nama_krt']}"
        if key in existing_keys:
            skip_count += 1
        else:
            new_records.append(rec)
    
    print(f"[INFO] Skip {skip_count} duplikat, Upload {len(new_records)} baru")
    
    if not new_records:
        print("[OK] Tidak ada data baru untuk diupload.")
        return 0
    
    # Batch upload (50 per batch)
    BATCH_SIZE = 50
    uploaded = 0
    errors = 0
    for i in range(0, len(new_records), BATCH_SIZE):
        batch = new_records[i:i+BATCH_SIZE]
        try:
            supabase.table('anomali_data').insert(batch).execute()
            uploaded += len(batch)
            print(f"  ✅ Uploaded batch {i//BATCH_SIZE + 1}: {uploaded}/{len(new_records)}")
        except Exception as e:
            errors += len(batch)
            print(f"  ❌ Error batch {i//BATCH_SIZE + 1}: {e}")
    
    print(f"\n✅ Selesai! {uploaded} record berhasil, {errors} error.")
    return uploaded


# ===================== MAIN =====================
def main():
    excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 
                              'Missing NIK 9999_1-1000.xlsx')
    
    if not os.path.exists(excel_path):
        print(f"ERROR: File tidak ditemukan: {excel_path}")
        sys.exit(1)
    
    print("=" * 60)
    print("  UPLOAD ANOMALI: Missing NIK Anggota Keluarga")
    print("=" * 60)
    
    # Load petugas map untuk enrichment
    petugas_map = load_petugas_map()
    
    # Baca Excel
    rows = read_missing_nik_excel(excel_path)
    
    # Proses ke format DB
    records = process_rows(rows, petugas_map)
    
    # Preview 2 record
    print("\n[PREVIEW] 2 record pertama:")
    for r in records[:2]:
        print(f"  - {r['kab_code']} | {r['nama_krt']} | {r['nama_petugas']}")
        print(f"    Catatan: {r['catatan'][:100]}...")
    
    # Upload
    supabase = load_supabase_config()
    upload_to_supabase(records, supabase)


if __name__ == '__main__':
    main()
