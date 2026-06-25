"""
laporan_progress_pencacah.py
============================
Script untuk membuat laporan progres per pencacah (petugas lapangan)
dari data granular_assignments.json yang sudah ada di disk.

Output:
  - Laporan_Progres_Pencacah_SE_Umum_<timestamp>.csv
  - Laporan_Progres_Pencacah_SE_UB_<timestamp>.csv
  - Laporan_Progres_Pencacah_<timestamp>.xlsx (jika openpyxl tersedia)

Jalankan: python3 laporan_progress_pencacah.py
"""

import json
import base64
import gzip
import csv
import os
from collections import defaultdict
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
GRANULAR_PATH = os.path.join(SCRIPT_DIR, "granular_assignments.json")

def load_granular():
    print(f"[INFO] Memuat {GRANULAR_PATH}...")
    with open(GRANULAR_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    compressed = data.get("compressed_data")
    if not compressed:
        raise ValueError("compressed_data tidak ditemukan di granular_assignments.json")
    raw = gzip.decompress(base64.b64decode(compressed)).decode("utf-8")
    payload = json.loads(raw)
    print(f"  Data terakhir diupdate: {payload.get('updated_at', 'N/A')}")
    return payload

def build_region_index(regions_list):
    idx = {}
    for i, r in enumerate(regions_list):
        idx[i] = {
            "kab_name": r[1] if len(r) > 1 else "-",
            "kec_name": r[3] if len(r) > 3 else "-",
        }
    return idx

def build_petugas_stats(targets, petugas_list, statuses_list, regions_list, survey_flag_filter):
    reg_index = build_region_index(regions_list)
    stats = defaultdict(lambda: {
        "username": "", "fullname": "",
        "kabupaten": set(), "kecamatan": set(),
        "total": 0, "submitted": 0, "approved": 0,
        "rejected": 0, "draft": 0, "open": 0,
    })
    for t in targets:
        if len(t) < 8:
            continue
        stat_idx    = t[3]
        pet_idx     = t[4]
        reg_idx     = t[5]
        survey_flag = t[7]
        if survey_flag != survey_flag_filter:
            continue
        if pet_idx < 0 or pet_idx >= len(petugas_list):
            continue
        username = petugas_list[pet_idx][0]
        fullname = petugas_list[pet_idx][1] if len(petugas_list[pet_idx]) > 1 else username
        status_str = statuses_list[stat_idx] if stat_idx < len(statuses_list) else "OPEN"
        su = status_str.upper()
        s = stats[username]
        s["username"] = username
        s["fullname"] = fullname
        s["total"] += 1
        reg = reg_index.get(reg_idx, {})
        kab = reg.get("kab_name", "-")
        kec = reg.get("kec_name", "-")
        if kab and kab != "-":
            s["kabupaten"].add(kab)
        if kec and kec != "-":
            s["kecamatan"].add(kec)
        if "SUBMITTED" in su or "APPROVED" in su or "REJECTED" in su or "REVOKED" in su:
            s["submitted"] += 1
        if "APPROVED" in su:
            s["approved"] += 1
        if "REJECTED" in su or "REVOKED" in su:
            s["rejected"] += 1
        if "DRAFT" in su:
            s["draft"] += 1
        if su == "OPEN":
            s["open"] += 1
    return stats

def save_csv(stats, filepath, survey_label):
    rows = []
    for username, s in stats.items():
        total = s["total"]
        submitted = s["submitted"]
        pct = round(submitted / total * 100, 2) if total > 0 else 0.0
        rows.append({
            "Username": username,
            "Nama Petugas": s["fullname"],
            "Kabupaten": " | ".join(sorted(s["kabupaten"])),
            "Kecamatan": " | ".join(sorted(s["kecamatan"])),
            "Total Target": total,
            "Sudah Submit": submitted,
            "Disetujui (Approved)": s["approved"],
            "Ditolak (Rejected)": s["rejected"],
            "Draft": s["draft"],
            "Open (Belum)": s["open"],
            "% Capaian": pct,
        })
    rows.sort(key=lambda x: (-x["% Capaian"], -x["Total Target"]))
    headers = ["Username", "Nama Petugas", "Kabupaten", "Kecamatan",
               "Total Target", "Sudah Submit", "Disetujui (Approved)",
               "Ditolak (Rejected)", "Draft", "Open (Belum)", "% Capaian"]
    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  [OK] [{survey_label}] CSV disimpan: {os.path.basename(filepath)} ({len(rows)} baris)")
    return rows

def save_excel(rows_umum, rows_ub, filepath):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [SKIP] openpyxl tidak tersedia.")
        return

    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_sheet(ws, rows, title):
        if not rows:
            return
        ws.title = title
        headers = list(rows[0].keys())
        hfill = PatternFill("solid", fgColor="1E3A5F")
        hfont = Font(bold=True, color="FFFFFF")
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for ri, row in enumerate(rows, 2):
            pct = row.get("% Capaian", 0)
            if pct >= 100:
                fill = PatternFill("solid", fgColor="C6EFCE")
            elif pct >= 75:
                fill = PatternFill("solid", fgColor="FFEB9C")
            elif pct >= 50:
                fill = PatternFill("solid", fgColor="FFCC99")
            elif pct > 0:
                fill = PatternFill("solid", fgColor="FFC7CE")
            else:
                fill = PatternFill("solid", fgColor="E0E0E0")
            for ci, h in enumerate(headers, 1):
                val = row.get(h, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = fill
                cell.border = border
                if isinstance(val, float):
                    cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="right")
                elif isinstance(val, int):
                    cell.alignment = Alignment(horizontal="center")
        for ci, h in enumerate(headers, 1):
            max_len = len(str(h))
            for row in rows[:200]:
                val = str(row.get(h, ""))
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 45)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    ws1 = wb.active
    write_sheet(ws1, rows_umum, "Pencacah SE Umum")
    ws2 = wb.create_sheet()
    write_sheet(ws2, rows_ub, "Pencacah SE UB")
    wb.save(filepath)
    print(f"  [OK] Excel disimpan: {os.path.basename(filepath)}")

def print_summary(stats, label, top_n=15):
    sorted_list = sorted(stats.values(), key=lambda x: x["submitted"], reverse=True)
    total_petugas = len(sorted_list)
    belum_submit  = sum(1 for x in sorted_list if x["submitted"] == 0)
    total_submit  = sum(x["submitted"] for x in sorted_list)
    total_target  = sum(x["total"] for x in sorted_list)
    pct_prov = round(total_submit / total_target * 100, 2) if total_target > 0 else 0

    print(f"\n{'='*68}")
    print(f"  RINGKASAN PROGRES PENCACAH [{label}]")
    print(f"{'='*68}")
    print(f"  Total Pencacah              : {total_petugas:,}")
    print(f"  Belum Submit Sama Sekali    : {belum_submit:,}")
    print(f"  Sudah Submit >= 1 target    : {total_petugas - belum_submit:,}")
    print(f"  Total Target                : {total_target:,}")
    print(f"  Total Submitted             : {total_submit:,}  ({pct_prov}%)")
    print(f"\n  TOP {top_n} PENCACAH (submitted terbanyak):")
    print(f"  {'No':<4} {'Username':<38} {'Target':>7} {'Submit':>7} {'%':>7}  Kabupaten")
    print(f"  {'-'*4} {'-'*38} {'-'*7} {'-'*7} {'-'*7}  {'-'*25}")
    for i, p in enumerate(sorted_list[:top_n], 1):
        pct = round(p["submitted"] / p["total"] * 100, 1) if p["total"] > 0 else 0
        kab = " / ".join(sorted(p["kabupaten"]))[:30]
        print(f"  {i:<4} {p['username']:<38} {p['total']:>7} {p['submitted']:>7} {pct:>6.1f}%  {kab}")
    print(f"\n  PENCACAH BELUM MULAI SAMA SEKALI (sample 10):")
    zero_list = [p for p in sorted_list if p["submitted"] == 0 and p["total"] > 0]
    for p in zero_list[:10]:
        kab = " / ".join(sorted(p["kabupaten"]))[:30]
        print(f"  {p['username']:<38} target={p['total']:>4}  [{kab}]")
    print(f"{'='*68}")

def main():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    payload       = load_granular()
    petugas_list  = payload.get("petugas", [])
    statuses_list = payload.get("statuses", [])
    regions_list  = payload.get("regions", [])
    targets       = payload.get("targets", [])
    print(f"[INFO] Total target: {len(targets):,}  |  Total petugas: {len(petugas_list):,}")

    # SE Umum
    print("\n[INFO] Menghitung progres SE Umum...")
    stats_umum = build_petugas_stats(targets, petugas_list, statuses_list, regions_list, 0)
    print_summary(stats_umum, "SE Umum")
    csv_umum = os.path.join(SCRIPT_DIR, f"Laporan_Progres_Pencacah_SE_Umum_{timestamp}.csv")
    rows_umum = save_csv(stats_umum, csv_umum, "SE Umum")

    # SE UB
    print("\n[INFO] Menghitung progres SE UB...")
    stats_ub = build_petugas_stats(targets, petugas_list, statuses_list, regions_list, 1)
    print_summary(stats_ub, "SE UB")
    csv_ub = os.path.join(SCRIPT_DIR, f"Laporan_Progres_Pencacah_SE_UB_{timestamp}.csv")
    rows_ub = save_csv(stats_ub, csv_ub, "SE UB")

    # Excel
    xlsx_path = os.path.join(SCRIPT_DIR, f"Laporan_Progres_Pencacah_{timestamp}.xlsx")
    save_excel(rows_umum, rows_ub, xlsx_path)
    print(f"\nSelesai! File tersimpan di: {SCRIPT_DIR}")

if __name__ == "__main__":
    main()
