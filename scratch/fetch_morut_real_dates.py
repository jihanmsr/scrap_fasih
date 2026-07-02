import asyncio
import json
import base64
import gzip
import os
import socket
from datetime import datetime, timezone, timedelta
from urllib.parse import unquote
from playwright.async_api import async_playwright
import pandas as pd
from openpyxl import load_workbook

# WITA timezone offset helper
def parse_iso_to_wita_date(iso_str):
    try:
        cleaned = iso_str.strip()
        if cleaned.endswith("Z"):
            cleaned = cleaned[:-1] + "+00:00"
        dt_utc = datetime.fromisoformat(cleaned)
        wita_offset = timezone(timedelta(hours=8))
        dt_wita = dt_utc.astimezone(wita_offset)
        return dt_wita.strftime("%Y-%m-%d")
    except Exception:
        return None

def check_port_open(port=9222):
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            return s.connect_ex(('localhost', port)) == 0
    except:
        return False

async def get_browser_context(p):
    port = 9223 if check_port_open(9223) else 9222
    if check_port_open(port):
        print(f"[INFO] Menghubungkan ke browser di port {port}...")
        try:
            browser = await p.chromium.connect_over_cdp(f"http://127.0.0.1:{port}")
            context = browser.contexts[0]
            page = None
            for p_page in context.pages:
                if "fasih-sm.bps.go.id" in p_page.url:
                    page = p_page
                    break
            if not page:
                page = await context.new_page()
                await page.goto("https://fasih-sm.bps.go.id/app/dashboard")
            return browser, context, page, True
        except Exception as e:
            print(f"[WARNING] Gagal connect CDP: {e}. Mencoba fallback...")
            
    # Fallback: launch persistent context (HEADLESS=FALSE agar user bisa login jika session mati)
    print("[INFO] Port CDP tertutup. Membuka browser Chrome secara visual...")
    try:
        abs_user_data_dir = os.path.abspath("playwright_chrome_profile")
        
        # Unlock profile if locked by a crashed process
        lock_file = os.path.join(abs_user_data_dir, "SingletonLock")
        if os.path.exists(lock_file) or os.path.islink(lock_file):
            try:
                os.unlink(lock_file)
                print("[INFO] Menghapus SingletonLock untuk membuka kunci profile.")
            except Exception as le:
                pass
                
        socket_file = os.path.join(abs_user_data_dir, "SingletonSocket")
        if os.path.exists(socket_file) or os.path.islink(socket_file):
            try:
                os.unlink(socket_file)
            except:
                pass
                
        chrome_path = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
        extra_args = ["--no-first-run", "--no-default-browser-check", "--disable-blink-features=AutomationControlled"]
        
        context = await p.chromium.launch_persistent_context(
            user_data_dir=abs_user_data_dir, headless=False, executable_path=chrome_path,
            ignore_default_args=["--enable-automation"],
            args=extra_args
        )
        page = context.pages[0] if context.pages else await context.new_page()
        print("[INFO] Membuka halaman dashboard FASIH...")
        await page.goto("https://fasih-sm.bps.go.id/app/dashboard", timeout=45000)
        
        # Cek jika dialihkan ke halaman login
        await asyncio.sleep(3)
        if "login" in page.url:
            print("\n==============================================================")
            print("[WARNING] Sesi FASIH Anda telah kedaluwarsa.")
            print("Harap LOGIN terlebih dahulu pada jendela browser Chrome yang baru terbuka.")
            print("Script akan mendeteksi otomatis jika Anda sudah berhasil login...")
            print("==============================================================\n")
            
            while "login" in page.url:
                await asyncio.sleep(2)
                
            print("[SUCCESS] Sesi berhasil dideteksi! Melanjutkan proses...")
            
        return None, context, page, False
    except Exception as e:
        print(f"[ERROR] Fallback gagal: {e}")
        print("\n====================================================================")
        print("[PERINGATAN] Chrome profile sedang dikunci/digunakan oleh aplikasi lain.")
        print("Harap TUTUP semua jendela Google Chrome Anda terlebih dahulu,")
        print("atau buka Chrome dengan debug mode: --remote-debugging-port=9222")
        print("====================================================================\n")
        return None, None, None, False

async def main():
    script_dir = "/Users/jihanmaisaroh/scrap_fasih"
    morut_json = os.path.join(script_dir, "granular_assignments_se_umum_7212.json")
    
    if not os.path.exists(morut_json):
        print(f"[ERROR] File granular {morut_json} tidak ditemukan!")
        return
        
    print(f"Membaca data Morowali Utara...")
    with open(morut_json, "r", encoding="utf-8") as f:
        data = json.load(f)
        
    comp = data.get("compressed_data")
    raw = json.loads(gzip.decompress(base64.b64decode(comp)).decode('utf-8'))
    
    statuses_list = raw.get("statuses", [])
    targets = raw.get("targets", [])
    
    # Filter target yang statusnya sudah Selesai (bukan OPEN dan DRAFT)
    non_selesais = {"OPEN", "DRAFT", "-", ""}
    completed_targets = []
    
    for t in targets:
        tid = t[0]
        stat_idx = t[3]
        status_str = statuses_list[stat_idx] if stat_idx < len(statuses_list) else "-"
        if status_str.upper() not in non_selesais:
            completed_targets.append(tid)
            
    total_completed = len(completed_targets)
    print(f"Ditemukan {len(targets)} total target. Jumlah target selesai: {total_completed}")
    
    if total_completed == 0:
        print("[WARNING] Tidak ada target selesai untuk dilacak tanggalnya.")
        return
        
    async with async_playwright() as p:
        browser, context, page, is_cdp = await get_browser_context(p)
        if not page:
            print("[ERROR] Gagal membuka browser. Batalkan proses.")
            return
            
        print("\n--- Memulai Penarikan Tanggal Submit Asli via Evaluasi Browser (Bypass Firewall) ---")
        
        date_counts = {}
        batch_size = 40 # Batch size optimal
        
        # Split completed_targets menjadi batches
        batches = [completed_targets[i:i + batch_size] for i in range(0, total_completed, batch_size)]
        
        print(f"Total batch: {len(batches)} (Ukuran batch: {batch_size})")
        
        # JS Fetch script yang dieksekusi langsung di context browser (melewati cookie & F5 firewall)
        js_fetch_script = """
            async (targetIds) => {
                const promises = targetIds.map(async (id) => {
                    try {
                        const res = await fetch(`/app/api/assignment-general/api/assignment-history/get-by-assignment-id?assignmentId=${id}`);
                        if (res.ok) {
                            const logs = await res.json();
                            return { id, logs, success: true };
                        }
                        return { id, error: `HTTP ${res.status}`, success: false };
                    } catch (e) {
                        return { id, error: e.toString(), success: false };
                    }
                });
                return await Promise.all(promises);
            }
        """
        
        processed_count = 0
        for b_idx, batch in enumerate(batches):
            try:
                # Cek jika dialihkan ke halaman login saat proses berjalan
                if "login" in page.url:
                    print("\n==============================================================")
                    print("[WARNING] Sesi FASIH Anda terputus mid-way!")
                    print("Harap LOGIN kembali pada jendela browser Chrome yang terbuka.")
                    print("Script akan mendeteksi otomatis jika Anda sudah berhasil login...")
                    print("==============================================================\n")
                    while "login" in page.url:
                        await asyncio.sleep(2)
                    print("[SUCCESS] Sesi berhasil dideteksi kembali! Melanjutkan...")
                    # Refresh page reference context after login
                    await asyncio.sleep(2)

                # Jalankan fetch di dalam browser page context
                results = await page.evaluate(js_fetch_script, batch)
                
                # Proses hasil batch di python
                for r in results:
                    processed_count += 1
                    if isinstance(r, dict) and r.get("success"):
                        history_list = r.get("logs", [])
                        
                        submit_date = None
                        if isinstance(history_list, list):
                            for log in history_list:
                                if isinstance(log, dict):
                                    status_name = str(log.get("assignmentStatusName", "")).upper()
                                    if "SUBMIT" in status_name or "APPROV" in status_name or "EDITED" in status_name:
                                        date_created = log.get("dateCreated")
                                        if date_created:
                                            submit_date = parse_iso_to_wita_date(date_created)
                                            
                        if not submit_date and isinstance(history_list, list) and len(history_list) > 0:
                            first_log = history_list[0]
                            if isinstance(first_log, dict):
                                date_created = first_log.get("dateCreated")
                                if date_created:
                                    submit_date = parse_iso_to_wita_date(date_created)
                                
                        if submit_date:
                            date_counts[submit_date] = date_counts.get(submit_date, 0) + 1
                    else:
                        pass # Abaikan yang gagal
                        
                if processed_count % 200 == 0 or processed_count == total_completed:
                    print(f"   [PROGRESS] Berhasil memproses {processed_count} / {total_completed} target...")
                    
            except Exception as e:
                print(f"[WARNING] Gagal memproses batch {b_idx}: {e}")
                # Jika page crash atau context hancur, coba recover page
                try:
                    page = context.pages[0] if context.pages else await context.new_page()
                except:
                    pass
                await asyncio.sleep(3)
                
        # Cleanup browser
        if browser:
            await browser.disconnect()
        else:
            await context.close()
            
    if not date_counts:
        print("[ERROR] Gagal menarik tanggal submit asli.")
        return
        
    # Susun dataframe
    sorted_dates = sorted(date_counts.keys())
    
    # Rekonstruksi data harian kumulatif
    smooth_history = []
    cumulative_sum = 0
    
    for date_str in sorted_dates:
        day_count = date_counts[date_str]
        cumulative_sum += day_count
        smooth_history.append({
            "Tanggal": date_str,
            "Total Selesai": cumulative_sum,
            "Submit Harian": day_count
        })
        
    df_history = pd.DataFrame(smooth_history)
    
    print("\n--- HASIL REAL PROGRESS HARIAN MOROWALI UTARA (BPS HISTORY) ---")
    print(df_history.to_string(index=False))
    
    # Save CSV
    csv_path = os.path.join(script_dir, "Morowali_Utara_Progres_Harian.csv")
    df_history.to_csv(csv_path, index=False)
    print(f"\n✅ CSV Progres Harian Asli disimpan ke {csv_path}")
    
    # Save Excel
    excel_path = os.path.join(script_dir, "Laporan_Morowali_Utara_7212.xlsx")
    if os.path.exists(excel_path):
        print(f"Menulis sheet Progres_Harian ke {excel_path}...")
        try:
            book = load_workbook(excel_path)
            if "Progres_Harian" in book.sheetnames:
                del book["Progres_Harian"]
            
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                df_history.to_excel(writer, sheet_name="Progres_Harian", index=False)
            print("✅ Excel Laporan Morowali Utara diperbarui dengan TANGGAL ASLI BPS!")
        except Exception as e:
            print(f"[ERROR] Gagal menulis ke Excel: {e}")

if __name__ == "__main__":
    asyncio.run(main())
