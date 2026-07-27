import openpyxl

wb = openpyxl.load_workbook('/Users/jihanmaisaroh/scrap_fasih/Progres Sulteng Fasih SM SE2026.xlsx', data_only=True)

# Data 26 Juli
kab_names = [
    "[01] BANGGAI KEPULAUAN", "[02] BANGGAI", "[03] MOROWALI", "[04] POSO",
    "[05] DONGGALA", "[06] TOLI-TOLI", "[07] BUOL", "[08] PARIGI MOUTONG",
    "[09] TOJO UNA-UNA", "[10] SIGI", "[11] BANGGAI LAUT", "[12] MOROWALI UTARA", "[71] PALU"
]
kab_codes = [7201, 7202, 7203, 7204, 7205, 7206, 7207, 7208, 7209, 7210, 7211, 7212, 7271]

data_26 = [31456, 93094, 32580, 65692, 76284, 53052, 36811, 112991, 41123, 67928, 20342, 24646, 90435]
# Let's dynamically pull from sheet just to be safe
ws_26 = wb['26 Juli']
data_26_dict = {}
for row in ws_26.iter_rows():
    if row[0].value in kab_names:
        # Col M (index 11) is Total, Col L (index 10 is Not Draft? Wait, looking at my previous dump)
        # Previous dump: ('[01] BANGGAI KEPULAUAN', None, 51722) -> This was col 11.
        pass

# The user already gave the table in the screenshot!
# The values for 26th are:
not_draft_26 = {
    7201: 31456,
    7202: 93094,
    7203: 32580,
    7204: 65692,
    7205: 76284,
    7206: 53052,
    7207: 36811,
    7208: 112991,
    7209: 41123,
    7210: 67928,
    7211: 20342,
    7212: 24646,
    7271: 90435
}
total_target_26 = {
    7201: 51722,
    7202: 163652,
    7203: 64828,
    7204: 117561,
    7205: 126134,
    7206: 93005,
    7207: 66237,
    7208: 192420,
    7209: 67038,
    7210: 119727,
    7211: 31809,
    7212: 47871,
    7271: 197640
}

# Values for 27th
not_draft_27 = {}
ws_27 = wb['27 Juli']
for row in ws_27.iter_rows():
    if row[0].value in kab_codes:
        # Based on previous dump, Not Draft + Open is row[-4]
        not_draft_27[row[0].value] = row[-4].value

# Calculate Differences (Trend)
differences = {}
total_diff = 0
for code in kab_codes:
    diff = not_draft_27[code] - not_draft_26[code]
    differences[code] = max(0, diff) # ensure no negative trend
    total_diff += max(0, diff)

# Calculate Adjustment
target_adjustment = 4761
adjusted_26 = {}
total_adjusted = 0

for code in kab_codes:
    if total_diff > 0:
        proportion = differences[code] / total_diff
        adjustment = round(proportion * target_adjustment)
    else:
        adjustment = round(target_adjustment / len(kab_codes)) # fallback evenly
        
    adjusted_value = not_draft_26[code] - adjustment
    adjusted_26[code] = adjusted_value
    total_adjusted += adjusted_value

# Handle rounding errors
diff_from_target = 741673 - total_adjusted
if diff_from_target != 0:
    # Add/sub remainder to the one with the biggest trend
    biggest_trend_code = max(differences, key=differences.get)
    adjusted_26[biggest_trend_code] += diff_from_target
    
print("--- HASIL ADJUSTMENT TANGGAL 26 JULI ---")
print(f"{'Kabupaten/Kota':<25} | {'Not Draft Asli':<15} | {'Pengurangan':<15} | {'Not Draft Baru':<15} | {'% Baru':<10}")
print("-" * 90)

total_baru = 0
total_target = 0
for code, name in zip(kab_codes, kab_names):
    asli = not_draft_26[code]
    baru = adjusted_26[code]
    pengurangan = asli - baru
    target = total_target_26[code]
    pct = (baru / target) * 100 if target > 0 else 0
    total_baru += baru
    total_target += target
    print(f"{name:<25} | {asli:<15} | -{pengurangan:<14} | {baru:<15} | {pct:.2f}%")

print("-" * 90)
print(f"{'TOTAL':<25} | {746434:<15} | -{746434-total_baru:<14} | {total_baru:<15} | {(total_baru/total_target)*100:.2f}%")
