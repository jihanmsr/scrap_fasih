import openpyxl

excel_path = "/Users/jihanmaisaroh/scrap_fasih/Progres Sulteng Fasih SM SE2026.xlsx"
wb = openpyxl.load_workbook(excel_path)
for name in wb.sheetnames:
    sheet = wb[name]
    print(f"Checking sheet: {name}")
    for r in range(1, sheet.max_row+1):
        for c in range(1, sheet.max_column+1):
            cell = sheet.cell(r, c)
            val = cell.value
            if val == '-' or (isinstance(val, str) and '-' in val and not val.startswith('=')):
                print(f"  Found at {cell.coordinate}: '{val}'")
