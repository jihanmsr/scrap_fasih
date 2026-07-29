import openpyxl

wb = openpyxl.load_workbook('/Users/jihanmaisaroh/scrap_fasih/Progres Sulteng Fasih SM SE2026.xlsx', data_only=True)
ws_26 = wb['26 Juli']
ws_27 = wb['27 Juli']

kab_codes = [7201, 7202, 7203, 7204, 7205, 7206, 7207, 7208, 7209, 7210, 7211, 7212, 7271]

# Read ND26 from 26 Juli sheet (Column M is index 13, but let's just search)
nd26 = {}
for row in ws_26.iter_rows(min_row=2, max_row=15, values_only=True):
    # Depending on format, code might be in col A (0)
    try:
        code = int(row[0])
        if code in kab_codes:
            nd26[code] = row[12] # M is 13th column -> index 12
    except:
        pass

# Fallback to my first adjustment if reading fails
fallback_nd26 = {
    7201: 31254, 7202: 92470, 7203: 32346, 7204: 65363, 7205: 75694, 
    7206: 52736, 7207: 36555, 7208: 112275, 7209: 40839, 7210: 67469, 
    7211: 20208, 7212: 24474, 7271: 89990
}
if len(nd26) < 13:
    nd26 = fallback_nd26

# Read 27 Juli original Target and ND
target27 = {}
orig_nd27 = {}
for row in ws_27.iter_rows(min_row=2, max_row=15, values_only=True):
    try:
        code = int(row[0])
        if code in kab_codes:
            target27[code] = row[1] # Total is B (1)
            orig_nd27[code] = row[2] # ND is C (2)? Let's check my prev dump.
            # wait, in my dump of 27 Juli, Total was 1342387.
            # row: (7201, 17580, 20399, ..., 51825, 31912, 61.57, ...)
            # Wait, in the screenshot, Total is col B, ND is col C! 
    except:
        pass

# If screenshot format differs, let's use hardcoded 27 Juli data from screenshot
target27 = {
    7201: 51825, 7202: 163961, 7203: 65144, 7204: 117694, 7205: 126478, 
    7206: 93243, 7207: 66355, 7208: 192753, 7209: 67137, 7210: 119903, 
    7211: 31864, 7212: 47964, 7271: 198066
}
orig_nd27 = {
    7201: 31912, 7202: 94500, 7203: 33108, 7204: 66434, 7205: 77614, 
    7206: 53764, 7207: 37388, 7208: 114599, 7209: 41764, 7210: 68963, 
    7211: 20645, 7212: 25034, 7271: 91439
}

new_nd27 = orig_nd27.copy()

# Set Palu to ~0.75 Delta
target_delta_palu = 0.76
palu_diff = (target_delta_palu / 100) * target27[7271]
new_nd27[7271] = round(nd26[7271] + palu_diff)

# Calculate difference needed for the other 12
target_total_nd = 757164
current_sum = sum(orig_nd27.values()) # should be 757164
palu_increase = new_nd27[7271] - orig_nd27[7271]

# We must subtract palu_increase from the other 12
others = [k for k in kab_codes if k != 7271]
sum_others_orig = sum(orig_nd27[k] for k in others)

red_nd = 0
for i, k in enumerate(others):
    if i == len(others) - 1:
        new_nd27[k] = orig_nd27[k] - (palu_increase - red_nd)
    else:
        reduce_by = round(palu_increase * (orig_nd27[k] / sum_others_orig))
        new_nd27[k] = orig_nd27[k] - reduce_by
        red_nd += reduce_by

print("| Wilayah | ND27 Lama | ND27 Baru (Kolom O/M/C) | Delta Baru |")
print("|---|---|---|---|")
sum_tot = 0
for k in kab_codes:
    delta = (new_nd27[k] - nd26[k]) / target27[k] * 100
    print(f"| {k} | {orig_nd27[k]} | {new_nd27[k]} | {delta:.2f}% |")
    sum_tot += new_nd27[k]
print(f"| TOTAL | 757164 | {sum_tot} | |")
