import openpyxl

excel_path = "/Users/jihanmaisaroh/scrap_fasih/Progres Sulteng Fasih SM SE2026.xlsx"
wb = openpyxl.load_workbook(excel_path)
sheet = wb['7 Juli']

print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")
for r in range(1, sheet.max_row + 1):
    vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    # Check if this row is not completely empty
    if any(v is not None for v in vals):
        print(f"Row {r:02d}: {vals}")
