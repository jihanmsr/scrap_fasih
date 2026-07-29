import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('/Users/jihanmaisaroh/scrap_fasih/Progres Sulteng Fasih SM SE2026.xlsx')
ws = wb['26 Juli']

# The adjusted values for M2 to M14
adjusted_values = {
    '[01] BANGGAI KEPULAUAN': 31254,
    '[02] BANGGAI': 92470,
    '[03] MOROWALI': 32346,
    '[04] POSO': 65363,
    '[05] DONGGALA': 75694,
    '[06] TOLI-TOLI': 52736,
    '[07] BUOL': 36555,
    '[08] PARIGI MOUTONG': 112275,
    '[09] TOJO UNA-UNA': 40839,
    '[10] SIGI': 67469,
    '[11] BANGGAI LAUT': 20208,
    '[12] MOROWALI UTARA': 24474,
    '[71] PALU': 89990
}

code_mapping = {
    '[01] BANGGAI KEPULAUAN': 7201,
    '[02] BANGGAI': 7202,
    '[03] MOROWALI': 7203,
    '[04] POSO': 7204,
    '[05] DONGGALA': 7205,
    '[06] TOLI-TOLI': 7206,
    '[07] BUOL': 7207,
    '[08] PARIGI MOUTONG': 7208,
    '[09] TOJO UNA-UNA': 7209,
    '[10] SIGI': 7210,
    '[11] BANGGAI LAUT': 7211,
    '[12] MOROWALI UTARA': 7212,
    '[71] PALU': 7271
}

# Apply to M2 to M14 and change A2 to A14
total_baru = 0
green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

row_idx_for_color = 0
for row_idx in range(2, 15):
    kab_name = ws[f'A{row_idx}'].value
    if kab_name in adjusted_values:
        baru = adjusted_values[kab_name]
        ws[f'M{row_idx}'].value = baru
        total_baru += baru
        
        ws[f'A{row_idx}'].value = code_mapping[kab_name]
        
        # alternating green color like in sheet 27 (usually odd rows have color)
        if row_idx % 2 == 0:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).fill = green_fill

# Update Total at M15 and A15
ws['M15'].value = total_baru
ws['A15'].value = 7200

# Save the workbook
wb.save('/Users/jihanmaisaroh/scrap_fasih/Progres Sulteng Fasih SM SE2026.xlsx')
print("Successfully updated Excel file.")
